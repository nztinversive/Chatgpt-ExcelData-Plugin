from flask import Flask, request, jsonify, send_from_directory, Response, send_file
import pandas as pd
import os
import json
from openpyxl import load_workbook
import io
import requests

app = Flask(__name__)

@app.route('/')
def index():
    return "Excel API Plugin is running."

@app.route('/.well-known/ai-plugin.json', methods=['GET'])
def send_ai_plugin_json():
    with open(os.path.join(os.getcwd(), '.well-known', 'ai-plugin.json'), 'r') as f:
        return Response(f.read(), content_type='application/json')

@app.route('/openapi.yaml', methods=['GET'])
def send_openapi_spec():
    with open(os.path.join(os.getcwd(), 'openapi.yaml'), 'r') as f:
        return Response(f.read(), content_type='text/yaml')

@app.route("/logo.png")
def logo():
    return send_file(os.path.join(os.getcwd(), "logo.png"))

@app.route('/api/data', methods=['GET'])
def get_data():
    # Remove the line below
    # sheet_name = request.args.get('sheet_name', 'Sheet1')

    try:
        # Request the Excel file from the API
        response = requests.get('https://excelapi.nztinversive.repl.co/output.xlsx', stream=True)
        response.raise_for_status()

        # Load the Excel file from the response
        workbook = load_workbook(io.BytesIO(response.content))

        # Get the first sheet name
        sheet_name = workbook.sheetnames[0]
        sheet = workbook[sheet_name]

        # Read the data from the sheet
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        # Return data as JSON
        return jsonify(data)
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)



