import os
from flask import Flask, jsonify, request
from flask_cors import CORS
from openpyxl import load_workbook
from flask import send_file

app = Flask(__name__)
CORS(app)

@app.route("/", methods=["GET"])
def index():
  return "API Online"

@app.route('/output.xlsx', methods=['GET'])
def get_output_excel():
    try:
        file_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'output.xlsx')
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return jsonify({'error': f"File not found: {file_path}"}), 404
        return send_file(file_path, download_name='output.xlsx', as_attachment=True)
    except Exception as e:
        print(f"Internal server error: {str(e)}")
        return jsonify({'error': f"Internal server error: {str(e)}"}), 500

@app.route('/api/data', methods=['GET'])
def get_data():
   

    try:
        # Load the Excel file
        workbook = load_workbook('output.xlsx')

        # Get the first sheet name
        sheet_name = workbook.sheetnames[0]
        sheet = workbook[sheet_name]

        # Read the data from the sheet
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        # Return data as JSON
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 8080)))
